from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import openpyxl
import requests

ETF_URLS = {
    "XLK": "https://www.ssga.com/us/en/intermediary/etfs/state-street-technology-select-sector-spdr-etf-xlk",
    "XLF": "https://www.ssga.com/us/en/intermediary/etfs/state-street-financial-select-sector-spdr-etf-xlf",
    "XLE": "https://www.ssga.com/us/en/intermediary/etfs/state-street-energy-select-sector-spdr-etf-xle",
    "XLV": "https://www.ssga.com/us/en/intermediary/etfs/state-street-health-care-select-sector-spdr-etf-xlv",
    "XLY": "https://www.ssga.com/us/en/intermediary/etfs/state-street-consumer-discretionary-select-sector-spdr-etf-xly",
    "XLP": "https://www.ssga.com/us/en/intermediary/etfs/state-street-consumer-staples-select-sector-spdr-etf-xlp",
    "XLI": "https://www.ssga.com/us/en/intermediary/etfs/state-street-industrial-select-sector-spdr-etf-xli",
    "XLB": "https://www.ssga.com/us/en/intermediary/etfs/state-street-materials-select-sector-spdr-etf-xlb",
    "XLRE": "https://www.ssga.com/us/en/intermediary/etfs/state-street-real-estate-select-sector-spdr-etf-xlre",
    "XLC": "https://www.ssga.com/us/en/intermediary/etfs/state-street-communication-services-select-sector-spdr-fund-xlc",
    "XLU": "https://www.ssga.com/us/en/intermediary/etfs/state-street-utilities-select-sector-spdr-fund-xlu",
}


def _download_holdings_workbook_url(etf: str) -> str:
    page_url = ETF_URLS[etf]
    html = requests.get(page_url, timeout=30, headers={"User-Agent": "Mozilla/5.0"}).text
    match = re.search(r'href="([^"]*holdings-daily-us-en-[^"]+\.xlsx)"', html, re.I)
    if not match:
        raise ValueError(f"Could not find holdings workbook link for {etf}")
    href = match.group(1)
    if href.startswith("http"):
        return href
    return f"https://www.ssga.com{href}"


def fetch_holdings(etf: str) -> list[dict[str, Any]]:
    workbook_url = _download_holdings_workbook_url(etf)
    response = requests.get(workbook_url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
    response.raise_for_status()
    workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    header_index = None
    for idx, row in enumerate(rows):
        normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if "ticker" in normalized and "name" in normalized:
            header_index = idx
            break
    if header_index is None:
        raise ValueError(f"Could not find holdings header row for {etf}")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    holdings: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        values = dict(zip(headers, row))
        ticker = values.get("Ticker") or values.get("Symbol")
        name = values.get("Name")
        if not ticker or not name:
            continue
        ticker = str(ticker).strip().upper()
        if ticker in {"-", "N/A"}:
            continue
        if not re.fullmatch(r"[A-Z.\-]+", ticker):
            continue
        holdings.append(
            {
                "etf": etf,
                "ticker": ticker,
                "name": str(name).strip(),
                "weight": values.get("Weight") or values.get("Index Weight"),
                "shares_held": values.get("Shares Held") or values.get("Shares Held:") or values.get("Shares"),
            }
        )
    return holdings


def fetch_all_holdings(etfs: list[str] | None = None) -> list[dict[str, Any]]:
    etfs = etfs or list(ETF_URLS.keys())
    all_holdings: list[dict[str, Any]] = []
    for etf in etfs:
        all_holdings.extend(fetch_holdings(etf))
    return all_holdings
